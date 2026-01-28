import logging
import queue
import threading
import multiprocessing
from pathlib import Path
from datetime import datetime

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QStackedWidget,
    QPushButton, QLabel, QProgressBar, QTreeWidget, QTreeWidgetItem,
    QFileDialog, QMessageBox, QFrame, QSpacerItem, QSizePolicy, QLineEdit
)
from PyQt6.QtCore import QThread, pyqtSignal, QTimer, Qt, QSize
from PyQt6.QtGui import QFont, QDragEnterEvent, QDropEvent, QIcon, QPixmap

from core.pdf_processor import run_processing_task
from config.app_config import config
from utils.logger import log_queue, add_output_logging, remove_handler





class DropZoneLabel(QLabel):
    """Custom label that accepts drag and drop"""
    file_dropped = pyqtSignal(str)
    
    def __init__(self, text=""):
        super().__init__(text)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setObjectName("dropZone")
        self.setMinimumHeight(200)
        self.setMaximumHeight(250)
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet("""
                QLabel#dropZone {
                    background-color: #f9fafb;
                    border: 2px dashed #111827;
                }
            """)
    
    def dragLeaveEvent(self, event):
        self.setStyleSheet("")
    
    def dropEvent(self, event: QDropEvent):
        self.setStyleSheet("")
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files and files[0].lower().endswith('.pdf'):
            self.file_dropped.emit(files[0])


class MainWindow(QMainWindow):
    """Clean minimal 3-screen PDF Separator"""
    
    def __init__(self):
        super().__init__()
        # Data
        self.selected_file = None
        self.selected_folder = None
        self.output_log_handler = None
        self.total_pages = 0
        self.files_created = 0
        self.processing_start_time = None
        
        # Multiprocessing handles
        self.processing_process = None
        self.process_queue = None
        self.process_stop_event = None
        self.status_timer = QTimer()
        self.status_timer.timeout.connect(self.check_process_status)
        
        self.init_ui()
        self.apply_theme()
        self.set_window_palette()
        
        # Log polling timer
        self.log_timer = QTimer()
        self.log_timer.timeout.connect(self.update_logs)
        self.log_timer.start(100)
    
    def set_window_palette(self):
        """Set lighter window frame colors"""
        from PyQt6.QtGui import QPalette, QColor
        palette = self.palette()
        # Set window background to lighter color
        palette.setColor(QPalette.ColorRole.Window, QColor(249, 250, 251))
        self.setPalette(palette)
    
    def show_message(self, title, message, icon_type="info"):
        """Show styled message dialog matching app theme"""
        msg = QMessageBox(self)
        msg.setWindowTitle(title)
        msg.setText(message)
        
        if icon_type == "question":
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            # Set No as default (highlighted/safe option)
            msg.setDefaultButton(QMessageBox.StandardButton.No)
        elif icon_type == "warning":
            msg.setIcon(QMessageBox.Icon.Warning)
        elif icon_type == "error":
            msg.setIcon(QMessageBox.Icon.Critical)
        else:
            msg.setIcon(QMessageBox.Icon.Information)
        
        # Apply custom styling with proper button differentiation
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #ffffff;
                font-family: 'Segoe UI';
            }
            QMessageBox QLabel {
                color: #111827;
                font-size: 13px;
                padding: 10px;
            }
            /* Primary button (Yes/OK) - Teal */
            QPushButton {
                background-color: #5abfac;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
                font-weight: 500;
                min-width: 70px;
                max-width: 100px;
            }
            QPushButton:hover {
                background-color: #4aa89a;
            }
            /* Secondary buttons (No/Cancel) - Gray */
            QPushButton:!default {
                background-color: #f3f4f6;
                color: #374151;
                border: 1px solid #d1d5db;
            }
            QPushButton:!default:hover {
                background-color: #e5e7eb;
                border-color: #9ca3af;
            }
        """)
        
        # Set lighter window frame for dialog
        from PyQt6.QtGui import QPalette, QColor
        palette = msg.palette()
        palette.setColor(QPalette.ColorRole.Window, QColor(249, 250, 251))
        palette.setColor(QPalette.ColorRole.WindowText, QColor(255, 255, 255))
        msg.setPalette(palette)
        
        if icon_type == "question":
            return msg.exec() == QMessageBox.StandardButton.Yes
        else:
            msg.exec()
            return True
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("LCA Case Separator")
        # Larger fixed-size window
        self.setFixedSize(675, 525)
        
        # Set window icon if available
        try:
            icon_path = Path(__file__).parent.parent.parent / "assets" / "icon.png"
            if icon_path.exists():
                self.setWindowIcon(QIcon(str(icon_path)))
        except:
            pass
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(25, 25, 25, 25)
        main_layout.setSpacing(15)
        central_widget.setLayout(main_layout)
        
        # Title section with description
        title_layout = QVBoxLayout()
        title_layout.setSpacing(5)
        
        title = QLabel("LCA Case Separator")
        title.setFont(QFont("Segoe UI", 18, QFont.Weight.DemiBold))
        title.setStyleSheet("color: #111827;")
        title_layout.addWidget(title)
        
        subtitle = QLabel("Automatically separate PDF files by case number using OCR")
        subtitle.setFont(QFont("Segoe UI", 10))
        subtitle.setStyleSheet("color: #6b7280;")
        title_layout.addWidget(subtitle)
        
        main_layout.addLayout(title_layout)
        
        # Content card - more compact
        self.content_card = QFrame()
        self.content_card.setObjectName("contentCard")
        main_layout.addWidget(self.content_card, 1)
        
        # Stacked widget
        self.stacked_widget = QStackedWidget()
        card_layout = QVBoxLayout()
        card_layout.setContentsMargins(30, 30, 30, 30)
        card_layout.setSpacing(15)
        self.content_card.setLayout(card_layout)
        card_layout.addWidget(self.stacked_widget)
        
        # Create screens
        self.create_upload_screen()
        self.create_output_screen()
        self.create_processing_screen()
        self.create_complete_screen()
        
        self.stacked_widget.setCurrentIndex(0)
    
    def create_upload_screen(self):
        """Screen 1: Compact upload"""
        screen = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        # Subtitle
        subtitle = QLabel("Drag and drop files to create a new project.")
        subtitle.setFont(QFont("Segoe UI", 11))
        subtitle.setStyleSheet("color: #6b7280;")
        layout.addWidget(subtitle)
        
        # Drop zone - more compact
        self.drop_zone = DropZoneLabel()
        self.drop_zone.file_dropped.connect(self.on_file_dropped)
        
        # Drop zone content
        drop_layout = QVBoxLayout()
        drop_layout.setSpacing(6)
        
        # Icon - custom image
        icon_label = QLabel()
        icon_path = Path(__file__).parent.parent.parent / "assets" / "pdf_logo.png"
        if icon_path.exists():
            pixmap = QPixmap(str(icon_path))
            # Scale to 40x40 while maintaining aspect ratio
            scaled_pixmap = pixmap.scaled(75, 75, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            icon_label.setPixmap(scaled_pixmap)
        else:
            # Fallback to emoji if image not found
            icon_label.setText("📄")
            icon_label.setFont(QFont("Arial", 40))
            icon_label.setStyleSheet("color: #9ca3af;")
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Text
        drop_text = QLabel("Upload a PDF file")
        drop_text.setFont(QFont("Segoe UI", 12))
        drop_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_text.setStyleSheet("color: #111827;")
        
        drop_layout.addStretch()
        drop_layout.addWidget(icon_label)
        drop_layout.addWidget(drop_text)
        
        # Browse link (put back inside drop zone for proper alignment)
        browse_layout = QHBoxLayout()
        browse_layout.setSpacing(5)
        browse_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        or_label = QLabel("or ")
        or_label.setFont(QFont("Segoe UI", 11))
        or_label.setStyleSheet("color: #9ca3af; border: none; background: transparent;")
        
        self.browse_btn = QPushButton("click to browse")
        self.browse_btn.setObjectName("linkButton")
        self.browse_btn.clicked.connect(self.select_file)
        self.browse_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        
        browse_layout.addStretch()
        browse_layout.addWidget(or_label)
        browse_layout.addWidget(self.browse_btn)
        browse_layout.addStretch()
        
        drop_layout.addLayout(browse_layout)
        drop_layout.addStretch()
        
        self.drop_zone.setLayout(drop_layout)
        layout.addWidget(self.drop_zone)
        
        # Pages label - more visible
        self.pages_label = QLabel("")
        self.pages_label.setFont(QFont("Segoe UI", 12, QFont.Weight.DemiBold))
        self.pages_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.pages_label.setStyleSheet("color: #111827; padding: 10px 10px 12px 0px;")
        layout.addWidget(self.pages_label)
        
        # Buttons row - compact
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.continue_btn = QPushButton("Continue")
        self.continue_btn.setObjectName("primaryButton")
        self.continue_btn.setMinimumWidth(100)
        self.continue_btn.setMaximumWidth(120)
        self.continue_btn.setEnabled(False)
        self.continue_btn.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))
        
        button_layout.addStretch()
        button_layout.addWidget(self.continue_btn)
        
        layout.addLayout(button_layout)
        screen.setLayout(layout)
        self.stacked_widget.addWidget(screen)

    def create_output_screen(self):
        """Screen 2: Output selection (Explicit Browse Button)"""
        screen = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(50, 10, 50, 20) # Reduced bottom margin
        layout.setSpacing(10)
        
        # Section 1: Selected File
        lbl1 = QLabel("Selected File")
        lbl1.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        lbl1.setStyleSheet("color: #111827;")
        layout.addWidget(lbl1)
        
        layout.addSpacing(10)
        
        lbl_sub = QLabel("Selected File:")
        lbl_sub.setFont(QFont("Segoe UI", 10))
        lbl_sub.setStyleSheet("color: #6b7280;")
        layout.addWidget(lbl_sub)
        
        self.file_info_label = QLabel("Not selected")
        self.file_info_label.setFont(QFont("Segoe UI", 11))
        self.file_info_label.setStyleSheet("color: #374151;")
        layout.addWidget(self.file_info_label)
        
        # Spacer
        layout.addSpacing(10)
        
        # Section 2: Output Location
        lbl2 = QLabel("Select Output Location")
        lbl2.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        lbl2.setStyleSheet("color: #111827;")
        layout.addWidget(lbl2)
        
        layout.addSpacing(10)
        
        # Row: Path Input + Browse Button
        row = QHBoxLayout()
        row.setSpacing(10)
        
        # Path Input (ReadOnly)
        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText("Select a folder...")
        self.path_input.setReadOnly(True)
        self.path_input.setFixedHeight(42)
        
        # Browse Button
        browse_btn = QPushButton("Browse")
        browse_btn.setObjectName("secondaryButton")
        browse_btn.setFixedHeight(42)
        browse_btn.setMinimumWidth(100)
        browse_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        browse_btn.clicked.connect(self.select_output_folder)
        
        row.addWidget(self.path_input, 1) # Expand input
        row.addWidget(browse_btn)
        
        layout.addLayout(row)
        
        layout.addStretch()
        
        # Buttons Row (Bottom)
        btn_row = QHBoxLayout()
        # Back Link (Bottom Left)
        back_btn = QPushButton("Back")
        back_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        back_btn.setStyleSheet("""
            QPushButton {
                border: none;
                background: transparent;
                color: #6b7280;
                font-size: 14px;
                text-align: left;
            }
            QPushButton:hover {
                color: #111827;
            }
        """)
        back_btn.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(0))
        btn_row.addWidget(back_btn)
        
        btn_row.addStretch()
        
        # Start Button (Bottom Right)
        self.start_btn = QPushButton("Start Processing")
        self.start_btn.setObjectName("primaryButton")
        self.start_btn.setMinimumWidth(150)
        self.start_btn.setFixedHeight(42)
        self.start_btn.setEnabled(False)
        self.start_btn.clicked.connect(self.start_processing)
        btn_row.addWidget(self.start_btn)
        
        layout.addLayout(btn_row)
        
        screen.setLayout(layout)
        self.stacked_widget.addWidget(screen)
    
    def create_processing_screen(self):
        """Screen 2: Compact processing view"""
        screen = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # Title - smaller
        # Title - smaller
        title_label = QLabel("Processing Document")
        title_label.setFont(QFont("Segoe UI", 12))
        title_label.setStyleSheet("color: #111827;")
        layout.addWidget(title_label)
        
        # Progress row - simple, flat
        progress_layout = QHBoxLayout()
        progress_layout.setSpacing(12)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("modernProgress")
        self.progress_bar.setMinimumHeight(6)
        self.progress_bar.setMaximumHeight(6)
        self.progress_bar.setTextVisible(False)
        
        self.progress_percent = QLabel("0%")
        self.progress_percent.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
        self.progress_percent.setStyleSheet("color: #111827; min-width: 35px;")
        
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.progress_percent)
        layout.addLayout(progress_layout)
        
        # Log tree - more compact
        self.log_tree = QTreeWidget()
        self.log_tree.setHeaderLabels(["Operation", "Status"])
        self.log_tree.setObjectName("logTree")
        # Set column widths
        self.log_tree.setColumnWidth(0, 400)  # Operation column
        # Status column auto-sized
        self.log_tree.setAlternatingRowColors(True)
        layout.addWidget(self.log_tree, 1)
        
        # Stop button
        stop_layout = QHBoxLayout()
        stop_layout.setAlignment(Qt.AlignmentFlag.AlignRight)
        
        self.stop_btn = QPushButton("Cancel Processing")
        self.stop_btn.setObjectName("secondaryButton")
        self.stop_btn.setMinimumWidth(140)
        self.stop_btn.clicked.connect(self.stop_processing)
        stop_layout.addWidget(self.stop_btn)
        
        layout.addLayout(stop_layout)
        
        screen.setLayout(layout)
        self.stacked_widget.addWidget(screen)
    
    def create_complete_screen(self):
        """Screen 3: Simple completion"""
        screen = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Checkmark - smaller
        checkmark = QLabel("✓")
        checkmark.setFont(QFont("Arial", 60, QFont.Weight.Bold))
        checkmark.setStyleSheet("color: #10b981;")
        checkmark.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(checkmark)
        
        # Title
        title = QLabel("Processing Complete")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.DemiBold))
        title.setStyleSheet("color: #111827;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Message
        self.success_message = QLabel("Your document has been separated into 0 files.")
        self.success_message.setFont(QFont("Segoe UI", 11))
        self.success_message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.success_message.setStyleSheet("color: #6b7280; margin: 10px 0;")
        layout.addWidget(self.success_message)
        
        layout.addSpacing(10)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        button_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.excel_btn = QPushButton("Download Summary")
        self.excel_btn.setObjectName("secondaryButton")
        self.excel_btn.clicked.connect(self.export_excel)
        self.excel_btn.setMinimumWidth(140)
        
        new_btn = QPushButton("New Project")
        new_btn.setObjectName("primaryButton")
        new_btn.clicked.connect(self.reset_to_upload)
        new_btn.setMinimumWidth(120)
        
        button_layout.addWidget(self.excel_btn)
        button_layout.addWidget(new_btn)
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        screen.setLayout(layout)
        self.stacked_widget.addWidget(screen)
    
    def apply_theme(self):
        """Apply clean minimal theme matching references"""
        stylesheet = """
            QMainWindow {
                background-color: #f9fafb;
            }
            
            QFrame#contentCard {
                background-color: #ffffff;
                border-radius: 8px;
                border: 1px solid #e5e7eb;
            }
            
            QFrame#dropZone {
                background-color: #fafafa;
                border: 2px dashed #d1d5db;
                border-radius: 6px;
            }
            
            QLabel#dropZone {
                background-color: #fafafa;
                border: 2px dashed #d1d5db;
                border-radius: 6px;
            }
            
            QPushButton#primaryButton {
                background-color: #5abfac;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 20px;
                font-size: 13px;
                font-weight: 500;
                font-family: 'Segoe UI';
            }
            
            QPushButton#primaryButton:hover {
                background-color: #4aa89a;
            }
            
            QPushButton#primaryButton:disabled {
                background-color: #e5e7eb;
                color: #9ca3af;
            }
            
            QPushButton#secondaryButton {
                background-color: #f3f4f6;
                color: #374151;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 8px 20px;
                font-size: 13px;
                font-weight: 500;
                font-family: 'Segoe UI';
            }
            
            QPushButton#secondaryButton:hover {
                background-color: #e5e7eb;
                border-color: #9ca3af;
            }
            
            QPushButton#linkButton {
                background-color: transparent;
                color: #000000;
                border: none;
                padding: 0;
                font-size: 11px;
                font-weight: 400;
                text-decoration: underline;
                font-family: 'Segoe UI';
            }
            
            QPushButton#linkButton:hover {
                text-decoration: underline;
                color: #4aa89a;
            }
            
            QLineEdit {
                background-color: white;
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                padding: 0 12px;
                color: #374151;
                font-family: 'Segoe UI';
                font-size: 13px;
                selection-background-color: #5abfac;
            }
            
            QFrame#dropZone {
                background-color: #fafafa;
                border: 2px dashed #d1d5db;
                border-radius: 6px;
            }
            
            QProgressBar#modernProgress {
                background-color: #e5e7eb;
                border: none;
                border-radius: 3px;
            }
            
            QProgressBar#modernProgress::chunk {
                background-color: #000000;
                border-radius: 3px;
            }
            
            QTreeWidget#logTree {
                background-color: #ffffff;
                alternate-background-color: #f9fafb;
                color: #1f2937;
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                font-family: 'Segoe UI';
                font-size: 11px;
                outline: none;
            }
            
            QTreeWidget#logTree::item {
                padding: 4px 8px;
                border: none;
            }
            
            QTreeWidget#logTree::item:selected {
                background-color: #f3f4f6;
                color: #111827;
            }
            
            QHeaderView::section {
                background-color: #fafafa;
                color: #6b7280;
                padding: 6px 8px;
                border: none;
                border-bottom: 1px solid #e5e7eb;
                font-weight: 600;
                font-size: 10px;
                font-family: 'Segoe UI';
            }
        """
        self.setStyleSheet(stylesheet)
    
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select PDF File", "", "PDF Files (*.pdf)"
        )
        if file_path:
            self.on_file_dropped(file_path)

    def select_output_folder(self):
        """Select output directory"""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.selected_folder = folder
            self.path_input.setText(folder)
            self.path_input.setCursorPosition(0) # Scroll to start
            self.start_btn.setEnabled(True)

    def on_file_dropped(self, file_path):
        self.selected_file = file_path
        logging.info(f"File selected: {file_path}")
        try:
            import fitz
            doc = fitz.open(file_path)
            self.total_pages = len(doc)
            doc.close()
            
            filename = Path(file_path).name
            self.pages_label.setText(f"📄 {filename} ({self.total_pages} pages)")
            self.pages_label.setStyleSheet("color: #111827; margin-top: 8px; margin-bottom: 8px;")
            
            # Enable continue button on first screen
            self.continue_btn.setEnabled(True)
            
            # Update info for next screen
            self.file_info_label.setText(f"{filename} ({self.total_pages} pages)")
            
        except Exception as e:
            self.pages_label.setText(f"⚠️ Error: {str(e)}")
            self.pages_label.setStyleSheet("color: #dc2626; margin-top: 8px; margin-bottom: 8px;")
            self.continue_btn.setEnabled(False)
    
    def stop_processing(self):
        """Stop processing with confirmation"""
        if self.show_message(
            "Cancel Processing",
            "Are you sure you want to cancel the current processing?",
            "question"
        ):
            logging.info("User requested to cancel processing")
            if self.process_stop_event:
                self.process_stop_event.set()
                
            self.stop_btn.setEnabled(False)
            self.stop_btn.setText("Canceling...")
    
    def start_processing(self):
        if not self.selected_file or not self.selected_folder:
            self.show_message("Error", "Please select file and output folder!", "warning")
            return
        
        self.stacked_widget.setCurrentIndex(2)
        self.processing_start_time = datetime.now()
        logging.info(f"Started processing file: {self.selected_file}")
        
        # Add output logging
        if self.output_log_handler:
            remove_handler(self.output_log_handler)
        self.output_log_handler = add_output_logging(self.selected_folder)
        
        self.stop_btn.setEnabled(True)
        self.stop_btn.setText("Cancel Processing")
        
        self.log_tree.clear()
        self.progress_bar.setValue(0)
        self.progress_percent.setText("0%")
        self.files_created = 0
        self.current_summary_data = [] # Data from processor
        
        cfg = config.data.get('processing', {})
        # Prepare Multiprocessing
        self.process_queue = multiprocessing.Queue()
        self.process_stop_event = multiprocessing.Event()
        
        # Use simple args that can be pickled
        args = (
            str(self.selected_file), 
            str(self.selected_folder), 
            cfg, 
            self.process_queue, 
            self.process_stop_event
        )
        
        self.processing_process = multiprocessing.Process(
            target=run_processing_task, 
            args=args
        )
        self.processing_process.start()
        
        # Start polling timer
        self.status_timer.start(100)
    
    def check_process_status(self):
        """Poll the multiprocessing queue for status updates"""
        try:
            while self.process_queue and not self.process_queue.empty():
                item = self.process_queue.get_nowait()
                if item:
                    event_type, progress, message = item
                    
                    # Handle specific signals
                    if event_type == 'complete':
                        self.status_timer.stop()
                        self.processing_process.join()
                        self.on_processing_complete()
                    elif event_type == 'summary_data':
                        self.current_summary_data = message
                    elif event_type == 'error':
                        self.status_timer.stop()
                        self.processing_process.join()
                        self.on_processing_error(message)
                    else:
                        # Forward progress/info to existing callback
                        self.processor_callback(event_type, progress, message)
        except Exception:
            pass
    
    def processor_callback(self, event_type, progress, message):
        if event_type == 'progress':
            percent = int(progress * 100)
            self.progress_bar.setValue(percent)
            self.progress_percent.setText(f"{percent}%")
            
            item = QTreeWidgetItem([message, "Processing"])
            self.log_tree.addTopLevelItem(item)
            self.log_tree.scrollToBottom()
            
        elif event_type == 'info':
            item = QTreeWidgetItem([message, "Done"])
            self.log_tree.addTopLevelItem(item)
            self.log_tree.scrollToBottom()
            
            if "Saved" in message:
                self.files_created += 1
    
    def on_processing_complete(self):
        self.progress_bar.setValue(100)
        self.progress_percent.setText("100%")
        self.success_message.setText(f"Your document has been separated into {self.files_created} files.")
        self.stacked_widget.setCurrentIndex(3)
    
    def on_processing_error(self, error_msg):
        QMessageBox.critical(self, "Error", f"An error occurred:\n{error_msg}")
        self.reset_to_upload()
    
    def export_excel(self):
        """Export summary with file save dialog (supports appending sheets)"""
        try:
            if not self.current_summary_data:
                self.show_message("Info", "No data to export.", "info")
                return

            # Open save file dialog
            default_name = "Processing_Report.xlsx"
            default_path = str(Path(self.selected_folder) / default_name)
            
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Report",
                default_path,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not save_path:  # User cancelled
                return
            
            from openpyxl import Workbook, load_workbook
            
            p_save = Path(save_path)
            input_name = Path(self.selected_file).stem
            sheet_title = input_name[:30] # Excel limit 31 chars
            
            if p_save.exists():
                try:
                    wb = load_workbook(p_save)
                    if sheet_title in wb.sheetnames:
                        # Handle duplicate sheet name
                        sheet_title = f"{sheet_title[:25]}_{datetime.now().strftime('%M%S')}"
                    ws = wb.create_sheet(title=sheet_title)
                except:
                    # Fallback if corrupt
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_title
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_title
            
            # Headers
            ws.append(["PDF File Generated Name", "PDF Start Location", "PDF End Location"])
            
            # Data
            for row in self.current_summary_data:
                ws.append(list(row))
                
            # Formatting
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            
            wb.save(save_path)
            
            logging.info(f"Summary exported to: {save_path}")
            self.show_message("Success", f"Report saved to:\n{save_path}")
        except Exception as e:
            logging.error(f"Export failed: {e}")
            self.show_message("Error", f"Failed to export:\n{str(e)}", "error")
    
    def reset_to_upload(self):
        # Cleanup logger
        if self.output_log_handler:
            remove_handler(self.output_log_handler)
            self.output_log_handler = None
            
        self.selected_file = None
        self.selected_folder = None
        if hasattr(self, 'path_input'):
            self.path_input.clear()
        self.total_pages = 0
        self.files_created = 0
        self.pages_label.setText("")
        self.continue_btn.setEnabled(False)
        self.start_btn.setEnabled(False)
        self.stacked_widget.setCurrentIndex(0)
    
    def update_logs(self):
        # Process pending log messages (max 50 to avoid freezing GUI)
        try:
            count = 0
            while count < 50:
                msg = log_queue.get_nowait()
                # Optionally process msg here if needed, currently we just drain
                count += 1
        except queue.Empty:
            pass
    
    def closeEvent(self, event):
        """Always show confirmation when closing the app"""
        is_processing = self.processing_process and self.processing_process.is_alive()
        
        if is_processing:
            confirmed = self.show_message(
                "Confirm Exit",
                "Processing is in progress. Are you sure you want to exit?",
                "question"
            )
        else:
            confirmed = self.show_message(
                "Confirm Exit",
                "Are you sure you want to quit?",
                "question"
            )
        
        if confirmed:
            if is_processing:
                # Signal stop
                if self.process_stop_event:
                    self.process_stop_event.set()
                # Wait briefly then forcefully terminate if needed
                self.processing_process.join(timeout=2.0)
                if self.processing_process.is_alive():
                    self.processing_process.terminate()
            event.accept()
        else:
            event.ignore()
